import selenium
from bs4 import BeautifulSoup
from selenium import webdriver
import time
from openpyxl import Workbook, load_workbook
from konlpy.tag import Kkma
import collections
import numpy as np
import matplotlib.pyplot as plt
from wordcloud import WordCloud
from os import path
from PIL import Image
import os

result_list = []


def get_data():
    driver = webdriver.Chrome("./chromedriver")
    for i in range(1, 11):
        driver.get("https://www1.president.go.kr/petitions/best?page="+str(i))
        soup = BeautifulSoup(driver.page_source, "html.parser")

        for i in soup.select("#cont_view > div.cs_area > div > div > div.board.text > div.b_list.category > div.bl_body > ul > li"):
            title = i.find("div", {"class": "bl_subject"}).text[5:].strip()
            result_list.append(title)
        time.sleep(0.1)

    driver.close()


def save_data_excel():
    write_workbook = Workbook()
    write_cell = write_workbook.active

    for i in range(1, len(result_list)+1):
        write_cell.cell(i, 1, result_list[i-1])

    write_workbook.save("bluehouse.xlsx")


def get_data_excel():
    read_workbook = load_workbook("./bluehouse.xlsx")
    read_cell = read_workbook.active

    result_list.clear()

    for i in range(1, 151):
        data = read_cell.cell(i, 1).value
        result_list.append(data)


kkma = Kkma()
get_data_excel()

list_temp = []

for data in result_list:
    list_temp = list_temp + kkma.nouns(data)  # 단어 추출

return_list = []

for data in list_temp:
    if len(data) > 1:
        return_list.append(data)

last_text = ""

for data in return_list:
    last_text = last_text+" "+data  # word cloud 를 위한 문자열로 변환

print(last_text)


# print(wordcloud.words_)

# plt.figure(figsize=(12, 12))
# plt.imshow(wordcloud, interpolation="bilinear")
# plt.axis("off")
# plt.show()

d = path.dirname(__file__) if "__file__" in locals() else os.getcwd()

mask = np.array(Image.open(path.join(d, 'korea.png')))

wordcloud = WordCloud(font_path="./font.ttf",
                      background_color="white", mask=mask,).generate(last_text)

plt.figure(figsize=(12, 12))
plt.imshow(wordcloud, interpolation="bilinear")
plt.axis("off")
plt.show()
